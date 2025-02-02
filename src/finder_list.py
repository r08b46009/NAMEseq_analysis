import pandas as pd
# import tqdm as tqdm
import re
print(pd.__version__)
import h5py
import argparse
import os
# import tables
import matplotlib.pyplot as plt
import os

import matplotlib.ticker as ticker
from matplotlib.ticker import ScalarFormatter, MultipleLocator
from openpyxl.utils import get_column_letter

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import numpy as np
import os
import argparse
def classify_motif(motif):
    if len(motif) == 8:  # 假设 dimer 是 8 个字符
        return "dimer"
    elif len(motif) == 13:  # 假设 trimer 是 13 个字符
        return "trimer"
    else:
        return "other"



# 讀取數據


# 定義合併函數
def merge_rows(group):


    # 初始化 A, T, C, G 的總和
    a_sum = 0
    t_sum = 0
    c_sum = 0
    g_sum = 0

    # 遍歷 group 的每一行，根據 Strand 計算互補核苷酸的總和
    for idx, row in group.iterrows():
        if row["Strand"] == '+':
            a_sum += row["A"]
            t_sum += row["T"]
            c_sum += row["C"]
            g_sum += row["G"]
        elif row["Strand"] == '-':
            # 反向互補計算
            a_sum += row["T"]  # A <-> T
            t_sum += row["A"]
            c_sum += row["G"]  # C <-> G
            g_sum += row["C"]

    return pd.Series({
        "Chromosome": group["Chromosome"].iloc[0],
        "Gene": ", ".join(sorted(set(group["Gene"]))),  # 合併基因名稱並去重
        "Motif": ", ".join(sorted(set(group["Motif"]))),  # 合併Motif
        "BED_Start": group["BED_Start"].iloc[0],
        "BED_End": group["BED_End"].iloc[0],
        "Motif Type": ", ".join(sorted(set(group["Motif Type"]))),  # 合併Motif Type
        # "Strand2": ", ".join(sorted(set(group["Strand"]))),  # 合併Strand
        "Position": list(group["Position"]),  # 保留Position列表
        # "Nucleotide": ", ".join(sorted(set(group["Nucleotide"]))),  # 合併Nucleotide
        # "Strand": ", ".join(sorted(set(group["Strand"]))),  # 合併Strand
        "A(+)T(-)": a_sum,  # 使用計算後的總和
        "T(+)A(-)": t_sum,
        "C(+)G(-)": c_sum,
        "G(+)C(-)": g_sum,
        "Sum_nm": group["Sum_nm"].sum(),
        "A->G Ratio": group["A->G Ratio"].sum(),
        "A->T/(A->G+A->T) Ratio": group["A->T/(A->G+A->T) Ratio"].sum()
    })

# 根據 BED_Start 和 BED_End 分組合併


def Hierarchical_sort(location_com):    # 進行排序，按照層級
    print(f"當前工作目錄是: {location_com}")
    file_path = location_com + "/comparison_results.csv"  # 修改為您的 CSV 檔案路徑
    data = pd.read_csv(file_path)
    sorted_data = data.sort_values(
        by=['Significant', 'Chromosome', 'Position'],
        ascending=[False, True, True]  # Significant: True > False, 其他按升序
    )

    # 保存排序後的數據到新的 CSV 文件
    output_file = location_com + "/sorted_comparison_results.csv"
    sorted_data.to_csv(output_file, index=False)

    print(f"排序完成，結果已保存至 {output_file}")



def draw(current_directory):

    # 載入測試資料檔案
    # current_directory = f"{current_directory[56:]}"

    print(f"當前工作目錄是: {current_directory}")
    # current_directory = current_directory.replace('/', '_')

    file_path = current_directory + "/comparison_results.csv"  # 修改為您的 CSV 檔案路徑
    data = pd.read_csv(file_path)

    # 過濾資料
    filtered_data = pd.DataFrame(data.dropna(subset=["Chromosome", "Position", "Significant"]))

    true_count = filtered_data[filtered_data["Significant"] == "True"].shape[0]
    false_count = filtered_data[filtered_data["Significant"] == "False"].shape[0]

    # 計算 True Rate
    true_rate = true_count / (true_count + false_count) if (true_count + false_count) > 0 else 0
    fixed_chromosomes = ["NC_003279.8", "NC_003280.10", "NC_003281.10", "NC_003282.8", "NC_003283.11","NC_003284.9"]#"NC_001328.1"]
    filtered_data = filtered_data[filtered_data["Chromosome"].isin(fixed_chromosomes)]
    # print("filtered_data",filtered_data)
    filtered_data["Chromosome"] = pd.Categorical(
        filtered_data["Chromosome"],
        categories=fixed_chromosomes,
        ordered=True
    )
    filtered_data["Chromosome_Code"] = filtered_data["Chromosome"].cat.codes

    # print("debug",filtered_data)
    filtered_data = filtered_data.sort_values("Chromosome")
    # print("filtered_data",filtered_data)
    chromosome_lengths = pd.Series({
    "NC_003279.8": 15072434,
    "NC_003280.10": 15279421,
    "NC_003281.10": 13783801,
    "NC_003282.8": 17493829,
    "NC_003283.11": 20924180,
    "NC_003284.9": 17718942,
    # "NC_001328.1": 13794
    })

    # print(chromosome_lengths)



    # 如果過濾後沒有數據，給出提示並退出
    if filtered_data.empty:
        print(f"沒有匹配的數據在染色體: {fixed_chromosomes}")
        return
    # 繪製散點圖
    plt.figure(figsize=(30, 30))

    for i, chrom in enumerate(fixed_chromosomes):
        if chrom in chromosome_lengths:
            chrom_length = chromosome_lengths[chrom]
            # 添加矩形，用axvline作為分隔
            plt.axhspan(
                ymin=0,
                ymax=chrom_length,
                xmin=i*0.2 - 0.05,
                xmax=i*0.2 + 0.05,
                facecolor="lightblue",
                alpha=0.3
            )
    true_data = filtered_data[filtered_data["Significant"] == "True"]
    false_data = filtered_data[filtered_data["Significant"] == "False"]

    # 先绘制 False (紫色) 的点
    plt.scatter(
        false_data["Chromosome_Code"],    # x 轴
        false_data["Position"],           # y 轴
        c="purple",                       # 紫色
        alpha=0.6,                        # 透明度
        edgecolor="k",
        s=1000                            # 点大小
    )

    # 再绘制 True (青色) 的点，覆盖紫色点
    plt.scatter(
        true_data["Chromosome_Code"],     # x 轴
        true_data["Position"],            # y 轴
        c="cyan",                         # 青色
        alpha=0.6,                        # 不透明
        edgecolor="k",
        marker="*",
        s=1200                            # 更大的点大小
    )


    # 添加數量標籤和 True Rate
    plt.text(
        1, 1,  # 圖的右上角位置
        f"True: {true_count}\nFalse: {false_count}\nTrue Rate: {true_rate:.2f}",
        fontsize=30,
        ha="right",
        va="top",
        transform=plt.gca().transAxes,  # 使用相對位置 (axes fraction)
        bbox=dict(facecolor="white", alpha=0.8, edgecolor="black")
    )


    # 標題和標籤
    plt.title(f"{current_directory}", fontsize=30)
    plt.gca().yaxis.set_major_locator(ticker.MultipleLocator(5000000))
    formatter = ScalarFormatter(useMathText=True)
    formatter.set_powerlimits((0, 0))  # 强制科学计数法
    plt.gca().yaxis.set_major_formatter(formatter)
    plt.xticks(
        range(len(fixed_chromosomes)),  # 固定的刻度位置
        fixed_chromosomes,  # 固定的名稱
        rotation=45,
        fontsize=30
    )

    plt.ylabel("Position", fontsize=30)
    plt.grid(alpha=0.5, linestyle="--")

    plt.tight_layout()
    output_file = f"{current_directory}.png"
    plt.savefig(output_file, dpi=300)  # dpi=300 可生成高解析度圖片
    # print("success",output_file)





# def process_csv_row(csv_row, gene_name, HS, chromosome, pos, motif_matches, hse_motif_with_6mA_count, A_to_G, A_to_T,A_to_T_A_to_G_sum):
#     # print(csv_row)
#     A_to_T_ratio = csv_row.iloc[10]  # Last column
#     A_to_G_ratio = csv_row.iloc[9]  # Second last column
#     sum_nm = csv_row.iloc[8]
#     dT = csv_row.iloc[7]
#     dG = csv_row.iloc[6]
#     dC = csv_row.iloc[5]
#     dA = csv_row.iloc[4]
#     print(A_to_T_ratio,A_to_G_ratio,sum_nm,dT,dG,dC,dA)

    

#     direc = csv_row.iloc[3]


#     if A_to_G_ratio != 0 and A_to_T_ratio != 0:
#         hse_motif_with_6mA_count += 1
#         motif_matches.append({
#             'HS': HS,
#             'Gene Name': gene_name,
#             'Chromosome': chromosome,
#             'Position': pos,
#             'direc': direc,
#             'A': dA,
#             'C': dC,
#             'G': dG,
#             'T': dT,
#             'Sum_nm': sum_nm,
#             'A->G Ratio': A_to_G_ratio,
#             'A->T/(A->G+A->T) Ratio': A_to_T_ratio
#     })
#     return A_to_G, A_to_T

def check_bed_and_csv_optimized(df,bed_df=None, chromosome_id=None, start_nucleotide=None, end_nucleotide=None, gene_name=None,HS=None, direction=None, output_dic = None, merged_results = None):
    # Initialize a dictionary to store the check results
    match_dict = {}
    df.columns = ["Chromosome", "Position", "Nucleotide", "Strand", "A", "C", "G", "T", "Sum_nm", "A->G Ratio", "A->T/(A->G+A->T) Ratio"]
    # df["A->T/(A->G+A->T) Ratio"] = (df["A->T/(A->G+A->T) Ratio"] / (
    #     df["A->G Ratio"] + df["A->T/(A->G+A->T) Ratio"]
    # )).replace(0, np.nan)
    df["A->T/(A->G+A->T) Ratio"] = np.where(
        df["Strand"] == "+",
        np.where(
            (df["T"] + df["G"]) == 0,  # 分母為 0 的情況
            0,  # 如果分母為 0，填入 0
            df["T"] / (df["T"] + df["G"])  # 正向計算
        ),
        np.where(
            (df["C"] + df["A"]) == 0,  # 分母為 0 的情況
            0,  # 如果分母為 0，填入 0
            df["A"] / (df["C"] + df["A"])  # 反向計算
        )
    )
# Filter the CSV based on input range and strand
    csv_filtered = df[(df["Chromosome"] == chromosome_id) &
                      (df["Position"] >= start_nucleotide) &
                      (df["Position"] <= end_nucleotide)]
                    #   (df["Strand"] == direction)]

    # csv_filtered = df
    for idx, csv_row in csv_filtered.iterrows():
        chromosome = csv_row["Chromosome"]
        nucleotide_position = csv_row["Position"]
        key = (chromosome, nucleotide_position)  # Tuple of chromosome and position as the key

        match_dict[key] = {'In Range': True, 'In BED': False}

    hse_motif_count = 0

    if bed_df is not None:
        bed_filtered = bed_df[(bed_df['chromosome'] == chromosome_id) &
                              (bed_df['start'] >= start_nucleotide) &
                              (bed_df['end'] <= end_nucleotide)]
        acc =0
        for _, bed_row in bed_filtered.iterrows():
            
            motif = bed_row["motif"]
            
            sub_matches = list(re.finditer(r"GAA|TTC", motif, re.IGNORECASE))  # 转换为列表
            # if sub_matches:  # 如果有匹配结果
            #     for match in sub_matches:
            #         print(f"Match: {match.group()}, Start: {match.start()}, End: {match.end()}")
            for sub_match in sub_matches:
                sub_start = bed_row["start"] + sub_match.start() + 1
                sub_end = bed_row["start"] + sub_match.end()
                # print(sub_start,sub_end)
                hse_motif_count += 1
                for pos in range(sub_start, sub_end+1):
                    key = (chromosome_id, pos)
                    match_dict[key] = {'In Range': True, 'In BED': True}

    # 匹配的結果存入列表



    # motif_matches = []
    # motif_matches_2 = []

      # 使用 dictionary 以确保唯一性

    for key, status in match_dict.items():
        chromosome, pos = key  # Unpack the chromosome and position
        if status['In Range'] and status['In BED']:
            print(chromosome,pos)


            csv_row = csv_filtered[(csv_filtered["Position"] == pos) ]
            # print("debug", csv_row)

            bed_row = bed_df[(bed_df['chromosome'] == chromosome) & (bed_df['start'] <= pos) & (bed_df['end'] >= pos)]
            # print("bed_row",bed_row)

            if not bed_row.empty:
                bed_start = bed_row.iloc[0]["start"]
                bed_end = bed_row.iloc[0]["end"]
                motif = bed_row.iloc[0]["motif"]
                motif_Type = bed_row.iloc[0]["mer"]
                unique_key = (chromosome, pos, motif)

                if not csv_row.empty and unique_key not in merged_results:
                    print('debug2',chromosome,gene_name,motif,bed_start,bed_end,motif_Type,csv_row)
                    merged_results[unique_key] = {
                        "Chromosome": chromosome,
                        "Gene": gene_name,
                        "Motif": motif,
                        "BED_Start": bed_start +1,
                        "BED_End": bed_end,
                        "Motif Type": motif_Type,
                        **csv_row.iloc[0].to_dict()
                    }

                    # A_to_G, A_to_T = process_csv_row(csv_row.iloc[0], gene_name, HS, chromosome, pos, motif_matches,hse_motif_with_6mA_count,A_to_G,A_to_T,A_to_T_A_to_G_sum)

    return merged_results

# List to store all results
def save_motif_results_to_csv(motif_results, output_file_path):
    if motif_results:
        # Convert the list of results into a DataFrame
        results_df = pd.DataFrame(motif_results)

        # Save to CSV
        try:
            results_df.to_csv(output_file_path, index=False)
            print(f"Results successfully saved to {output_file_path}")
        except Exception as e:
            print(f"Error saving file: {e}")
    else:
        print("No results to save.")


data2 = pd.read_csv("/staging/biology/u8411596/NAME/output/preprocessed_data_926_HS_07_2_k40_2/N2_HS_NT_L3_adenine.csv", sep=',', index_col=False, header=None)
data = pd.read_csv("/staging/biology/u8411596/NAME/output/preprocessed_data_926_k40_2/N2_NT_L3_adenine.csv", sep=',', index_col=False, header=None)



# Now you can work with `data` and `data2` as pandas DataFrames

def main(output_dir):
    # Convert to a pandas DataFrame


    os.makedirs(output_dir, exist_ok=True)
    # gene_list_path = args.input_promoter_region
    gene_list_path = "/staging/biology/u8411596/NAME/1203/promoter_gene_ranges_within_" +range_nt +"_genes_" +name + ".csv"
    print(gene_list_path[-11:-4])
    # hse_output_file_path_2 = f'{output_dir}/hse_motif_ratio_results_2_{range_nt}_{gene_list_path[-11:-4]}_{mer}.csv'
    # save_motif_results_to_csv(hse_result_2, hse_output_file_path_2)


    print("1")
    bed_file_path = '/staging/biology/u8411596/NAME/motif_positions_with_type_new_2.bed'

    renew_dir =  output_dir + "/" + ver + "/dimer" + "/" + range_nt + "/" + name
    os.makedirs(renew_dir, exist_ok=True)
    bed_data = pd.read_csv(bed_file_path, sep=r'\s+', header=None, names=["chromosome", "start", "end", "motif", "mer"])
    print("2")



    print(gene_list_path)
    gene_list = pd.read_csv(gene_list_path)  # Ensure this is loaded before iterating
    print("3")
    # Iterate over each gene in the gene list
    merged_results_1 = {}
    merged_results_2 = {}
    for idx, gene_row in gene_list.iterrows():
        # print(idx)
        chromosome_id = gene_row['Chromosome (NC_*)']
        start_nucleotide = gene_row['Nucleotide Location(start)']
        end_nucleotide = gene_row['Nucleotide Location(end)']
        gene_name = gene_row['Gene']
        direc = gene_row['Strand']
        # Run the function to find matches and check A->T and A->G ratios
        # If you don't want to use the BED filter, pass None for bed_df
        output_file_match_dict = f'{renew_dir}/match_{range_nt}_{name}_{out_data1}_{mer}.csv'
        merged_results_1 = check_bed_and_csv_optimized(data, bed_df=bed_data, chromosome_id=chromosome_id, start_nucleotide=start_nucleotide, end_nucleotide=end_nucleotide,gene_name=gene_name,HS='non-HS',direction = direc, output_dic = output_file_match_dict, merged_results = merged_results_1)

        merged_results_2= check_bed_and_csv_optimized(data2, bed_df=bed_data, chromosome_id=chromosome_id, start_nucleotide=start_nucleotide, end_nucleotide=end_nucleotide,gene_name=gene_name,HS='HS',direction = direc, output_dic = output_file_match_dict, merged_results = merged_results_2)




    final_df = pd.DataFrame(merged_results_1.values())

    # 保存合并结果
    output_file = f"{renew_dir}/Adenine&Thymine_nonHS.csv"
    final_df.to_csv(output_file, index=False)
    print(f"合并结果已保存到 {output_file}")

    final_df = pd.DataFrame(merged_results_2.values())

    # 保存合并结果
    output_file = f"{renew_dir}/Adenine&Thymine_HS.csv"
    final_df.to_csv(output_file, index=False)
    print(f"合并结果已保存到 {output_file}")

 
    return renew_dir



def accumulated(raw_dir):
    print(raw_dir)
    print(renew_dir)
    df_sorted = pd.read_csv(f'{renew_dir}/Adenine&Thymine_HS.csv')  # replace with your HS file path
    
    df_sorted = df_sorted.groupby(["Chromosome", "BED_Start", "BED_End"]).apply(merge_rows).reset_index(drop=True)

    

    # 輸出結果
    output_file = f"{raw_dir}/Sites_HS.csv"  # 替換為你希望保存的文件名
    df_sorted.to_csv(output_file, index=False)
    print(f"合併完成，結果已保存到 {output_file}")

    print("success")

 
    df_sorted = pd.read_csv(f'{renew_dir}/Adenine&Thymine_nonHS.csv')  # replace with your HS file path
    # df_sorted = pd.read_csv(f'{renew_dir}/merged_results_2.csv')  # replace with your HS file path
    
    df_sorted = df_sorted.groupby(["Chromosome", "BED_Start", "BED_End"]).apply(merge_rows).reset_index(drop=True)


    df_sorted.to_csv(f"{raw_dir}/Sites_nonHS.csv", index=False)



def p_value(renew_dir,name_gene):
    file1_path = f'{renew_dir}/Sites_nonHS.csv'
    file2_path = f'{renew_dir}/Sites_HS.csv'

    data1 = pd.read_csv(file1_path)
    data2 = pd.read_csv(file2_path)

    # Merge the datasets on shared columns for analysis
    # print("data1",data1,"data2",data2)
    merged = pd.merge(data1, data2, on=["Chromosome","Gene", "Motif Type", "BED_Start", "BED_End"], suffixes=("_non_HS", "_HS"))

    # Calculate the ratio of Ratio2 values in the second dataset to the first dataset
    # merged["Ratio2_Division"] = merged["Ratio2_2"] / merged["Ratio2_1"]
    merged["A->T/(A->G+A->T) Ratio_HS/non_HS"] = merged["A->T/(A->G+A->T) Ratio_HS"] / merged["A->T/(A->G+A->T) Ratio_non_HS"]

    # print(merged)
    merged["A->T/(A->G+A->T) Ratio_HS/non_HS"] = merged["A->T/(A->G+A->T) Ratio_HS/non_HS"].replace([np.inf, -np.inf], np.nan)

    # 篩選非 NaN 的數據
    filtered_ratio2_division = merged["A->T/(A->G+A->T) Ratio_HS/non_HS"].fillna(0)

    # 對於大於 0 的值取 log2，小於等於 0 的值設為 0
    logs_value = np.where(filtered_ratio2_division > 0, np.log2(filtered_ratio2_division), 0)

    # 將結果存回 DataFrame
    merged["log_2"] = logs_value
    # signs = np.sign(logs_value)
    # print("sign",signs)
    # labels = np.where(signs > 0, "upregulated", "downregulated")
    # merged["signs"] = signs
    merged["log_2"] = logs_value

    merged["sort_flag"] = 0
    merged.loc[(merged["A->T/(A->G+A->T) Ratio_non_HS"] == 0) & (merged["A->T/(A->G+A->T) Ratio_HS"] != 0), "sort_flag"] = 1  # 優先級 1
    # merged.loc[(merged["Ratio2_1"] != 0) & (merged["Ratio2_2"] != 0), "sort_flag"] = 2  # 優先級 2
    merged.loc[(merged["A->T/(A->G+A->T) Ratio_non_HS"] != 0) & (merged["A->T/(A->G+A->T) Ratio_HS"] == 0), "sort_flag"] = 4  # 優先級 3
    merged.loc[(merged["A->T/(A->G+A->T) Ratio_non_HS"] == 0) & (merged["A->T/(A->G+A->T) Ratio_HS"] == 0), "sort_flag"] = 5  # 優先級 4



    # 根據 sort_flag 進行排序
    # merged_sorted = merged.sort_values(by="sort_flag").drop(columns="sort_flag")
    merged_sorted = merged.sort_values(
        by=["sort_flag", "A->T/(A->G+A->T) Ratio_HS/non_HS", "A->T/(A->G+A->T) Ratio_HS"],
        ascending=[True, False, False]  # 第一層升序，第二和第三層降序
    )#.drop(columns=["sort_flag"])  # 移除輔助欄位


    output_file = f"{renew_dir}/{name_gene}_colored_output.xlsx"

    # 建立 Excel Workbook 和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    merged_sorted.loc[(merged_sorted["A->T/(A->G+A->T) Ratio_HS/non_HS"]!= 0) & (merged_sorted["log_2"] > 0), "sort_flag"] = 2  # 優先級 5

    merged_sorted.loc[(merged_sorted["A->T/(A->G+A->T) Ratio_HS/non_HS"]!= 0) & (merged_sorted["log_2"] < 0), "sort_flag"] = 3  # 優先級 5

    # 定義顏色填充
    color_mapping = {
        1: PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),  # 柔和紅色
        2: PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # 淡紅色
        3: PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid"),  # 淡藍色
        4: PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid"),  # 柔和藍色
        # 4: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),  # 淺灰色
        5: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # 淺淺灰色
    }
    # 寫入標題
    merged_sorted = merged_sorted.sort_values(
        by=["sort_flag", "A->T/(A->G+A->T) Ratio_HS/non_HS", "A->T/(A->G+A->T) Ratio_HS", "sort_flag"],
        ascending=[True, False, False,True]  # 第一層升序，第二和第三層降序
    )#.drop(columns=["sort_flag"])  # 移除輔助欄位

    for col_num, header in enumerate(merged_sorted.columns, 1):
        ws.cell(row=1, column=col_num, value=header)

    # 寫入數據並根據 sort_flag 設置顏色
    # for i, row in enumerate(merged_sorted.itertuples(), start=2):  # 從 Excel 的第2行開始寫入數據
    #     for col_num, value in enumerate(row[1:], start=1):  # 跳過索引列
    #         cell = ws.cell(row=i, column=col_num, value=value)
    #         if col_num == len(row) - 1:  # 根據 sort_flag 塗整行顏色
    #             sort_flag = row.sort_flag
    #             for j in range(1, len(merged_sorted.columns) + 1):
    #                 ws.cell(row=i, column=j).fill = color_mapping.get(sort_flag, None)
    default_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")  # 預設白色

    for i, row in enumerate(merged_sorted.itertuples(), start=2):  # 從 Excel 第2行開始寫入數據
        sort_flag = row.sort_flag  # 假設 DataFrame 的 "sort_flag" 列存在
        fill = color_mapping.get(sort_flag, default_fill)  # 確保有預設填充樣式

        for col_num, value in enumerate(row[1:], start=1):  # 跳過索引列
            cell = ws.cell(row=i, column=col_num, value=value)
            # 只需為該行的單元格設置一次顏色
            cell.fill = fill
    for col_num, col in enumerate(merged_sorted.columns, start=1):
        column_letter = get_column_letter(col_num)  # 將數字轉換成 Excel 列字母
        max_length = max(merged_sorted[col].astype(str).map(len).max(), len(col)) + 2  # 計算最大字串長度，+2 以保留間距
        ws.column_dimensions[column_letter].width = max_length
    print(wb)
    wb.save(output_file)
    print(f"Excel 文件已保存為 {output_file}")

    # merged_sorted.to_csv(f"{renew_dir}/P-value.csv", index=False)



if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process CSV and BED files with output directory.")
    parser.add_argument('-o', '--output_dir', type=str, default='output', help="Specify the output directory for saving results.")
    parser.add_argument('-i', '--input_promoter_region', type=str, default='input', help="Specify the input directory for saving results.")
    parser.add_argument('-m', '--mer', type=str, default='m', help="Specify the input directory for saving results.")
    parser.add_argument('-l', '--length', type=str, default='l', help="Specify the input directory for saving results.")



    args = parser.parse_args()
    # parser = argparse.ArgumentParser(description="Process CSV and BED files with output directory.")
    vers = ["T<=1","old"]
    # ver = "old"

    # ###
    # name = args.input_promoter_region[69:-4]
    # mer = args.mer
    # range_nt = args.length

    # print("range_nt",range_nt)
    # print("mer",mer)
    # print("name",name)
    # names = ["lab_new"]
    # names = ["lab_new","Richter"]
    names = [ "random2"]
    # mers = ["trimer", "dimer"]

    # ranges = ["1500","2000"]
    # names = ["Richter"]
    # names = ["lab_new"] 
    # ranges = ["2000"]
    mers = ["dimer"]
    ranges = ["2000"]
    output_d = "output_0126"
    for v in vers:
        ver = v
        data2_name= f"/staging/biology/u8411596/NAME/output/preprocessed_data_1027_HS_{ver}/N2_HS_1_L6_adenine.csv"
        data1_name= f"/staging/biology/u8411596/NAME/output/preprocessed_data_1016_woHS_2_{ver}/N2_1_L6_adenine.csv"

        # data2 = pd.read_csv(data2_name, sep=',', index_col=False, header=None)
        # data = pd.read_csv(data1_name, sep=',', index_col=False, header=None)

        out_data2 = data2_name[56:-26]
        out_data1 = data1_name[56:-27]

        for i in names:
            name = i
            for j in mers:
                mer = j
                for k in ranges:
                    range_nt = k
                    # renew_dir = main(output_d)renew_dir =  output_d + "/" + ver + "/"+ mer + "/" + range_nt + "/" + name
                    renew_dir =  output_d + "/" + ver + "/"+ mer + "/" + range_nt + "/" + name

                    accumulated(renew_dir)
                    renew_dir =  output_d + "/" + ver + "/"+ mer + "/" + range_nt + "/" + name
                    # accumulated(renew_dir)
                    p_value(renew_dir,name)
                    # try:
                    # significant(mer,renew_dir)
                    # draw(renew_dir)
                    # Hierarchical_sort(renew_dir)
                    # except:
                    #     pass